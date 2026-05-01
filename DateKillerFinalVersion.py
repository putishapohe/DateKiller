# Time: 20250708 20:58
# Author: Putishapohe
# Location: shanghai 
# Version 8 Update
# 1. 添加代码注释
# 2. 优化代码结构
#------------------------------------#

import tkinter.messagebox
from tkinter import *
import tkinter as tk
from tkinter import filedialog
from tkinter import ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.pyplot as plt
import math
import numpy as np
from scipy.fftpack import fft
import xlsxwriter
import openpyxl
import threading
from queue import Queue, Empty

# =====================================================================================
# --- 子窗口与对话框类 (Sub-windows and Dialog Classes) ---
# =====================================================================================

class Adc_transfer(tk.Toplevel):
    """ADC参数设置子窗口，用于将原始数据转换为真实电压值。"""
    def __init__(self):
        super().__init__()
        self.title('ADC参数设置')
        self.geometry('200x60')
        self.resizable(False, False)
        self.userinfo = None  # 用于存储用户输入的结果
        self.setup_UI()

    def setup_UI(self):
        # --- 界面布局 ---
        tk.Label(self, text="ADC位数:").place(relx=0.1, rely=0.1)
        tk.Label(self, text="基准电压:").place(relx=0.1, rely=0.51)
    
        self.var = tk.StringVar()
        tk.Entry(self, textvariable=self.var, width=5).place(relx=0.5, rely=0.1)
    
        self.adc = tk.StringVar()
        tk.Entry(self, width=5, textvariable=self.adc).place(relx=0.5, rely=0.51)
    
        tk.Button(self, text='确定', command=self.ok).place(relx=0.8, rely=0.0)
        tk.Button(self, text='退出', command=self.cancel).place(relx=0.8, rely=0.51)
  
    def ok(self):
        """确认按钮回调，保存用户输入并关闭窗口。"""
        try:
            self.userinfo = [int(self.var.get()), float(self.adc.get())]
            self.destroy()
        except ValueError:
            tkinter.messagebox.showerror("错误", "请输入有效的数字。")

    def cancel(self):
        """取消按钮回调，销毁窗口。"""
        self.userinfo = None
        self.destroy()

class digital_filter(tk.Toplevel):
    """数字滑动窗口滤波器设置子窗口。"""
    def __init__(self):
        super().__init__()
        self.title('数字滑动窗口滤波器')
        self.geometry('220x60')
        self.resizable(False, False)
        self.filterinfo = None
        self.setup_UI()
  
    def setup_UI(self):
        tk.Label(self, text="请输入滤波窗口宽度:").place(relx=0, rely=0.3)
    
        self.filterwidth_var = tk.StringVar()
        tk.Entry(self, textvariable=self.filterwidth_var, width=5).place(relx=0.6, rely=0.3)
    
        tk.Button(self, text='确定', command=self.ok).place(relx=0.85, rely=0.0)
        tk.Button(self, text='退出', command=self.cancel).place(relx=0.85, rely=0.51)
  
    def ok(self):
        """确认按钮回调，保存窗口宽度并关闭。"""
        try:
            self.filterinfo = [int(self.filterwidth_var.get())]
            self.destroy()
        except ValueError:
            tkinter.messagebox.showerror("错误", "请输入一个整数。")

    def cancel(self):
        """取消按钮回调，销毁窗口。"""
        self.filterinfo = None
        self.destroy()

class FFTSettingDialog(tk.Toplevel):
    """FFT参数设置对话框。"""
    def __init__(self, parent, default_fs=1000):
        super().__init__(parent)
        self.title("FFT SETTING")
        self.geometry("350x220")
        self.resizable(False, False)
        self.settings = None
        self.default_fs = default_fs
        self.setup_UI()

    def setup_UI(self):
        # --- 显示模式 (线性/分贝) ---
        mode_frame = ttk.LabelFrame(self, text="显示模式")
        mode_frame.pack(padx=10, pady=5, fill="x")
        self.fft_mode_var = tk.StringVar(value="Linear")
        ttk.Radiobutton(mode_frame, text="线性坐标", variable=self.fft_mode_var, value="Linear").pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(mode_frame, text="分贝(dB)", variable=self.fft_mode_var, value="dB").pack(side=tk.LEFT, padx=5)

        # --- 窗函数选择 ---
        window_frame = ttk.LabelFrame(self, text="窗函数")
        window_frame.pack(padx=10, pady=5, fill="x")
        self.window_var = tk.StringVar(value="None (Rectangular)")
        window_options = ["None (Rectangular)", "Hanning", "Hamming", "Blackman"]
        ttk.Combobox(window_frame, textvariable=self.window_var, values=window_options, state="readonly").pack(pady=5, padx=5, fill="x")

        # --- 采样频率设置 (自动/手动) ---
        fs_frame = ttk.LabelFrame(self, text="采样频率 (Fs)")
        fs_frame.pack(padx=10, pady=5, fill="x")
        self.fs_mode_var = tk.StringVar(value="Auto")
        ttk.Radiobutton(fs_frame, text="自动计算", variable=self.fs_mode_var, value="Auto", command=self.toggle_fs_entry).pack(side=tk.LEFT, padx=5)
        ttk.Radiobutton(fs_frame, text="手动输入 (Hz):", variable=self.fs_mode_var, value="Manual", command=self.toggle_fs_entry).pack(side=tk.LEFT, padx=5)
        self.fs_value_var = tk.StringVar(value=str(self.default_fs))
        self.fs_entry = ttk.Entry(fs_frame, textvariable=self.fs_value_var, width=10)
        self.fs_entry.pack(side=tk.LEFT, padx=5)
    
        # --- 确认/取消按钮 ---
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=10)
        tk.Button(btn_frame, text="确认", command=self.confirm).pack(side=tk.LEFT, padx=10)
        tk.Button(btn_frame, text="取消", command=self.cancel).pack(side=tk.LEFT, padx=10)

        self.toggle_fs_entry()  # 根据默认选项设置输入框初始状态

    def toggle_fs_entry(self):
        """根据选择的采样率模式，启用或禁用手动输入框。"""
        self.fs_entry.config(state="normal" if self.fs_mode_var.get() == "Manual" else "disabled")

    def confirm(self):
        """收集所有设置，验证后保存并关闭窗口。"""
        fs_value = 0
        if self.fs_mode_var.get() == "Manual":
            try:
                fs_value = float(self.fs_value_var.get())
                if fs_value <= 0:
                    tkinter.messagebox.showerror("错误", "采样频率必须为正数")
                    return
            except ValueError:
                tkinter.messagebox.showerror("错误", "请输入有效的采样频率")
                return
    
        self.settings = {
            "fft_mode": self.fft_mode_var.get(),
            "window": self.window_var.get(),
            "fs_mode": self.fs_mode_var.get(),
            "fs_value": fs_value
        }
        self.destroy()

    def cancel(self):
        self.settings = None
        self.destroy()

class DataGeneratorDialog(tk.Toplevel):
    """数据生成器对话框，用于创建测试数据。"""
    def __init__(self, parent):
        super().__init__(parent)
        self.parent = parent
        self.title("创建数据")
        self.geometry("400x450")
        self.resizable(False, False)
        self.generated_data = None
        self.setup_ui()

    def setup_ui(self):
        # --- 数据类型选择 ---
        type_frame = ttk.LabelFrame(self, text="1. 选择数据类型")
        type_frame.pack(padx=10, pady=5, fill="x")
        self.data_type_var = tk.StringVar(value="含噪正弦波")
        self.data_types = ["单频正弦波", "多频正弦波", "高斯白噪声", "含噪正弦波"]
        type_combo = ttk.Combobox(type_frame, textvariable=self.data_type_var, values=self.data_types, state="readonly")
        type_combo.pack(pady=5, padx=5, fill="x")
        type_combo.bind("<<ComboboxSelected>>", self.toggle_parameter_frames)

        # --- 通用参数 ---
        common_frame = ttk.LabelFrame(self, text="2. 通用参数")
        common_frame.pack(padx=10, pady=5, fill="x")
        ttk.Label(common_frame, text="采样点数:").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.points_var = tk.StringVar(value="4096")
        ttk.Entry(common_frame, textvariable=self.points_var, width=15).grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(common_frame, text="采样率 (Hz):").grid(row=1, column=0, padx=5, pady=2, sticky="w")
        self.fs_var = tk.StringVar(value="10000")
        ttk.Entry(common_frame, textvariable=self.fs_var, width=15).grid(row=1, column=1, padx=5, pady=2)

        # --- 信号参数 ---
        self.signal_frame = ttk.LabelFrame(self, text="3. 信号参数")
        self.signal_frame.pack(padx=10, pady=5, fill="x")
        ttk.Label(self.signal_frame, text="频率 (Hz):").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.freq_var = tk.StringVar(value="50")
        ttk.Entry(self.signal_frame, textvariable=self.freq_var, width=25).grid(row=0, column=1, padx=5, pady=2)
        ttk.Label(self.signal_frame, text="(多频用英文逗号隔开)").grid(row=1, column=1, padx=5, pady=0, sticky="w")
        ttk.Label(self.signal_frame, text="振幅 (V):").grid(row=2, column=0, padx=5, pady=2, sticky="w")
        self.amp_var = tk.StringVar(value="1.0")
        ttk.Entry(self.signal_frame, textvariable=self.amp_var, width=25).grid(row=2, column=1, padx=5, pady=2)
        ttk.Label(self.signal_frame, text="(与频率一一对应)").grid(row=3, column=1, padx=5, pady=0, sticky="w")

        # --- 噪声参数 ---
        self.noise_frame = ttk.LabelFrame(self, text="4. 噪声参数")
        self.noise_frame.pack(padx=10, pady=5, fill="x")
        ttk.Label(self.noise_frame, text="噪声标准差 (σ):").grid(row=0, column=0, padx=5, pady=2, sticky="w")
        self.noise_var = tk.StringVar(value="0.1")
        ttk.Entry(self.noise_frame, textvariable=self.noise_var, width=15).grid(row=0, column=1, padx=5, pady=2)

        # --- 按钮 ---
        btn_frame = tk.Frame(self)
        btn_frame.pack(pady=15)
        ttk.Button(btn_frame, text="生成", command=self.generate).pack(side=tk.LEFT, padx=20)
        ttk.Button(btn_frame, text="取消", command=self.destroy).pack(side=tk.LEFT, padx=20)

        self.toggle_parameter_frames()

    def toggle_parameter_frames(self, event=None):
        """根据所选数据类型，动态显示或隐藏参数设置区域。"""
        data_type = self.data_type_var.get()
        # 控制信号参数区域的可见性
        signal_state = 'normal' if data_type in ["单频正弦波", "多频正弦波", "含噪正弦波"] else 'disabled'
        for child in self.signal_frame.winfo_children():
            child.configure(state=signal_state)
        # 控制噪声参数区域的可见性
        noise_state = 'normal' if data_type in ["高斯白噪声", "含噪正弦波"] else 'disabled'
        for child in self.noise_frame.winfo_children():
            child.configure(state=noise_state)

    def generate(self):
        """根据用户设置的参数生成数据。"""
        try:
            # 解析通用参数
            num_points = int(self.points_var.get())
            fs = float(self.fs_var.get())
            if num_points <= 0 or fs <= 0:
                raise ValueError("采样点数和采样率必须为正")

            time_axis = np.linspace(0, num_points / fs, num_points, endpoint=False)
            final_signal = np.zeros(num_points)
            data_type = self.data_type_var.get()

            # 生成信号部分
            if data_type in ["单频正弦波", "多频正弦波", "含噪正弦波"]:
                freqs_str = self.freq_var.get().split(',')
                amps_str = self.amp_var.get().split(',')
                if len(freqs_str) != len(amps_str):
                    raise ValueError("频率和振幅的数量必须一致")
                freqs = [float(f.strip()) for f in freqs_str]
                amps = [float(a.strip()) for a in amps_str]
                for freq, amp in zip(freqs, amps):
                    final_signal += amp * np.sin(2 * np.pi * freq * time_axis)

            # 生成噪声部分
            if data_type in ["高斯白噪声", "含噪正弦波"]:
                noise_sigma = float(self.noise_var.get())
                final_signal += noise_sigma * np.random.randn(num_points)

            self.generated_data = (time_axis, final_signal)
            self.destroy()

        except ValueError as e:
            tkinter.messagebox.showerror("参数错误", f"输入参数无效: {e}")
        except Exception as e:
            tkinter.messagebox.showerror("生成失败", f"发生未知错误: {e}")

# =====================================================================================
# --- 主应用界面类 (Main Application Class) ---
# =====================================================================================

class MY_GUI(tk.Tk):
    """主应用程序类，整合了所有UI和功能。"""
    def __init__(self):
        super().__init__()
        self.title("DataKiller")
        self.geometry('760x500')
        self.resizable(False, False)
    
        # --- 初始化核心变量 ---
        self.filename = ""
        self.columns = ["Time", "Data"]
        self.num = 0
        self.refv = 0
        self.linesnum = 0
    
        # 初始化数据容器，确保程序启动即可进行编辑操作
        self.time_con = np.array([])
        self.data_con = np.array([])
        self.vaue = []
        self.datafit = []

        # 初始化分析结果的存储变量
        self.Noise_Vpp_ex = ""
        self.Noise_3σ_ex = ""
        self.Noise_6σ_ex = ""
        self.function_fit_export = ""
    
        self.set_init_window()
  
    def set_init_window(self):
        """设置主窗口的全部UI组件。"""
    
        # --- 分析结果显示区域 ---
        self.fitR_lable = Label(self, text="线性度 R=")
        self.fitR_lable.place(relx=0.75, rely=0.80)
        self.fit_label = Label(self, text="拟合公式 :")
        self.fit_label.place(relx=0.42, rely=0.8)
        self.RMS_label = Label(self, text="Vpp_Noise=")
        self.RMS_label.place(relx=0.42, rely=0.87)
        self.RMS_label = Label(self, text="3σ_Noise=")
        self.RMS_label.place(relx=0.63, rely=0.87)
        self.Authour_label = Label(self, text="联系作者：putishapohe")
        self.Authour_label.place(relx=0.8, rely=0.95)
    
        self.fit_Text = Text(self, width=23, height=1)
        self.fit_Text.place(relx=0.50, rely=0.81)
        self.fitR_Text = Text(self, width=8, height=1)
        self.fitR_Text.place(relx=0.83, rely=0.81)
        self.Max_Text = Text(self, width=8, height=1)
        self.Max_Text.place(relx=0.53, rely=0.88)
        self.Ave_Text = Text(self, width=8, height=1)
        self.Ave_Text.place(relx=0.73, rely=0.88)
    
        # --- 功能按钮 ---
        self.button_fit = tk.Button(self, text="线性拟合", command=self.fitresult)
        self.button_fit.place(relx=0.33, rely=0.8)
        self.button_RMS = tk.Button(self, text="RMS噪声", command=self.RMSCAL)
        self.button_RMS.place(relx=0.33, rely=0.87)
        self.button_draw = tk.Button(self, text="绘制曲线", command=self.draw)
        self.button_draw.place(relx=0.33, rely=0.73)
    
        # --- 菜单栏设置 ---
        menubar = Menu(self, tearoff=False)
    
        # 文件菜单
        self.filemenu = Menu(menubar, tearoff=False)
        self.filemenu.add_command(label='新建...', command=self.open_data_generator)
        self.filemenu.add_command(label='打开...', command=self.import_data)
        self.filemenu.add_separator()
        self.filemenu.add_command(label='保存', command=self.export_data)
        self.filemenu.add_separator()
        self.filemenu.add_command(label='退出', command=self.quit)
        menubar.add_cascade(label='文件(W)', menu=self.filemenu)
    
        # 数据菜单
        datamenu = Menu(menubar, tearoff=False)
        menubar.add_cascade(label="数据(S)", menu=datamenu)
        datamenu.add_command(label="真实值转化", command=self.setup_config)
        datamenu.add_command(label="FFT 设置与分析", command=self.open_fft_settings)
        datamenu.add_separator()
        datamenu.add_command(label="区间选择", command=self.regionchoose)
        datamenu.add_separator()
        datamenu.add_command(label="数字滤波", command=self.datafilter_config)
        datamenu.add_separator()
        datamenu.add_command(label="全部清除", command=self.clear_all_data)
    
        # 设置与帮助菜单
        optionmenu = Menu(menubar, tearoff=False)
        menubar.add_cascade(label="设置(S)", menu=optionmenu)
        optionmenu.add_command(label="点赞", command=self.option_Danm)
        optionmenu.add_command(label="吐槽", command=self.option_complain)
        helpmenu = Menu(menubar, tearoff=False)
        menubar.add_cascade(label="帮助(H)", menu=helpmenu)
        helpmenu.add_command(label="使用文档", command=self.help_instrument)
        helpmenu.add_command(label="技术支持", command=self.help_support)
        self.config(menu=menubar)
    
        # --- 数据表格与控制按钮 ---
        table_frame = ttk.Frame(self)
        table_frame.place(relx=0.004, rely=0.028, relwidth=0.308, relheight=0.958)

        tree_container = ttk.Frame(table_frame)
        tree_container.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_container, show="headings", columns=self.columns)
        self.tree.column("Time", width=80, anchor="center")
        self.tree.column("Data", width=80, anchor="center")
        self.tree.heading("Time", text="时间")
        self.tree.heading("Data", text="数据")
    
        self.scrollbar = ttk.Scrollbar(tree_container, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=self.scrollbar.set)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tree.bind("<Double-1>", self.on_tree_double_click) # 绑定双击事件

        button_frame = ttk.Frame(table_frame)
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=5)
        ttk.Button(button_frame, text="添加行", command=self.add_table_row, width=8).pack(side=tk.LEFT, padx=2, expand=True)
        ttk.Button(button_frame, text="删除行", command=self.delete_table_row, width=8).pack(side=tk.LEFT, padx=2, expand=True)
        ttk.Button(button_frame, text="粘贴", command=self.paste_from_clipboard, width=8).pack(side=tk.LEFT, padx=2, expand=True)
    
        # --- 绘图区域 ---
        self.f = Figure(figsize=(6, 4), dpi=80)
        self.a = self.f.add_subplot(111)
        self.a.grid(True)
        self.a.set_xlabel("Time(s)")
        self.a.set_ylabel("Voltage")
        self.a.set_title("ADC-Data")
        self.canvas = FigureCanvasTkAgg(self.f, self)
        self.canvas.draw()
        self.canvas.get_tk_widget().place(relx=0.33, rely=0.028)
    
        # --- Matplotlib工具栏 ---
        self.toolbar = NavigationToolbar2Tk(self.canvas, self)
        self.toolbar.update()
        self.toolbar.place(relx=0.5, rely=0.69)

        # --- 绑定鼠标事件 ---
        self.canvas.mpl_connect('scroll_event', self.on_scroll)
        self.canvas.mpl_connect('button_press_event', self.on_pan_press)
        self.canvas.mpl_connect('motion_notify_event', self.on_pan_motion)
        self.canvas.mpl_connect('button_release_event', self.on_pan_release)
  
    # ----------------------------------------------------------------------------------
    # --- 帮助与提示信息 (Help and Info Messages) ---
    # ----------------------------------------------------------------------------------
  
    def help_instrument(self):
        """显示“使用文档”信息。"""
        tkinter.messagebox.showinfo(title='oOPs', message='Hi~你好，朋友，没有使用文档，因为懒得写 ￣ω￣=')
  
    def help_support(self):
        """显示“技术支持”信息。"""
        tkinter.messagebox.showinfo(title='技术支持', message='请联系 L00570201')
  
    def option_Danm(self):
        """显示“点赞”信息。"""
        tkinter.messagebox.showinfo(title='谢谢', message='谢谢支持！(((o(*ﾟ▽ﾟ*)o)))♡')
  
    def option_complain(self):
        """显示“吐槽”信息。"""
        tkinter.messagebox.showwarning(title='oOPs', message='使用次数已耗尽，请自行点击退出，不送 ￣へ￣')
  
    def export_erro(self):
        """显示导出错误信息。"""
        tkinter.messagebox.showwarning(title='截取数据异常', message='请先选择数据区间')

    # ----------------------------------------------------------------------------------
    # --- 核心功能方法 (Core Feature Methods) ---
    # ----------------------------------------------------------------------------------

    def clear_all_data(self):
        """清除所有数据、图表和分析结果，重置UI状态。"""
        # 1. 清空Treeview中的所有项目
        self.tree.delete(*self.tree.get_children())

        # 2. 重置所有内部数据变量为空
        self.time_con = np.array([])
        self.data_con = np.array([])
        self.vaue = []
        self.datafit = []
        self.linesnum = 0
        self.Noise_Vpp_ex = ""
        self.Noise_3σ_ex = ""
        self.Noise_6σ_ex = ""
        self.function_fit_export = ""
    
        # 3. 重置全局状态标志为默认值
        global stat
        stat = 0

        # 4. 清空Matplotlib图表并重置标题和坐标轴
        self.a.clear()
        self.a.grid(True)
        self.a.set_xlabel("Time(s)")
        self.a.set_ylabel("Voltage")
        self.a.set_title("ADC-Data")
        self.canvas.draw_idle()

        # 5. 清空所有分析结果文本框
        self.fit_Text.delete(1.0, tk.END)
        self.fitR_Text.delete(1.0, tk.END)
        self.Max_Text.delete(1.0, tk.END)
        self.Ave_Text.delete(1.0, tk.END)
  
    def open_data_generator(self):
        """打开数据生成器对话框并处理返回的数据。"""
        dialog = DataGeneratorDialog(self)
        self.wait_window(dialog)  # 等待对话框关闭

        # 如果用户生成了数据 (而不是直接关闭窗口)
        if dialog.generated_data is not None:
            time_data, value_data = dialog.generated_data
            self.update_data_variables(time_data, value_data)

            # 在UI上更新数据
            self.tree.delete(*self.tree.get_children())
            for item in self.vaue:
                self.tree.insert('', tk.END, values=item)
        
            self.draw() # 绘制新生成的数据
        
            global stat
            stat = 0
            tkinter.messagebox.showinfo("成功", f"成功生成 {self.linesnum} 条数据")
  
    def import_data(self):
        """打开文件对话框，并使用后台线程导入数据以避免UI冻结。"""
        file_path = filedialog.askopenfilename(
            title="选择数据文件",
            filetypes=[("Excel files", "*.xlsx"), ("Text files", "*.txt"), ("All files", "*.*")])
        if not file_path:
            return

        # 创建一个队列用于线程间通信
        self.queue = Queue()
        # 创建并启动后台线程，daemon=True表示主线程退出时子线程也退出
        self.thread = threading.Thread(target=self._load_data_worker, args=(file_path, self.queue), daemon=True)
        self.thread.start()

        # 禁用“打开”按钮防止重复操作，并更新窗口标题提示用户
        self.filemenu.entryconfig("打开...", state="disabled")
        self.title("DataKiller - 正在导入数据...")

        # 启动队列检查器，以轮询方式从后台线程获取结果
        self.after(100, self._process_queue)

    def _load_data_worker(self, file_path, queue):
        """在后台线程中运行，负责读取和解析文件，并将结果放入队列。"""
        try:
            if file_path.endswith('.txt'):
                time_data, value_data = self._load_txt_data_internal(file_path)
            elif file_path.endswith('.xlsx'):
                time_data, value_data = self._load_xlsx_data_internal(file_path)
            else:
                # 将错误信息放入队列
                queue.put({"status": "error", "message": "文件类型不支持，请选择 .txt 或 .xlsx 文件"})
                return
        
            # 将成功的结果和数据放入队列
            queue.put({"status": "success", "data": (time_data, value_data)})

        except Exception as e:
            # 将异常信息放入队列
            queue.put({"status": "error", "message": f"无法读取文件: {e}"})

    def _process_queue(self):
        """在主线程中运行，检查后台线程的结果并安全地更新UI。"""
        try:
            # 非阻塞地从队列获取消息
            result = self.queue.get_nowait()

            # 无论成功失败，都恢复UI状态
            self.filemenu.entryconfig("打开...", state="normal")
            self.title("DataKiller")

            if result["status"] == "success":
                time_data, value_data = result["data"]
                self.update_data_variables(time_data, value_data)
            
                # 更新UI
                self.tree.delete(*self.tree.get_children())
                for item in self.vaue:
                    self.tree.insert('', tk.END, values=item)
            
                global stat
                stat = 0
                tkinter.messagebox.showinfo("成功", f"成功导入 {self.linesnum} 条数据")
        
            elif result["status"] == "error":
                tkinter.messagebox.showerror("导入错误", result["message"])

        except Empty:
            # 如果队列为空，说明后台线程仍在工作，100ms后再次检查
            self.after(100, self._process_queue)

    def _load_txt_data_internal(self, file_path):
        """(内部使用) 从txt文件加载数据，返回时间和数据列表。"""
        time_data, value_data = [], []
        with open(file_path, 'r') as f:
            for line in f:
                line = line.strip()
                if not line: continue
                # 支持制表符或空格作为分隔符
                parts = line.replace('\t', ' ').split()
                if len(parts) >= 2:
                    time_data.append(float(parts[0]))
                    value_data.append(float(parts[1]))
        return time_data, value_data

    def _load_xlsx_data_internal(self, file_path):
        """(内部使用) 从xlsx文件加载数据，返回时间和数据列表。"""
        time_data, value_data = [], []
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=1, max_col=2, values_only=True):
            if row[0] is not None and row[1] is not None:
                try:
                    time_data.append(float(row[0]))
                    value_data.append(float(row[1]))
                except (ValueError, TypeError):
                    # 如果行内容无法转换为数字，则跳过
                    continue
        return time_data, value_data

    def update_data_variables(self, time_data, value_data):
        """更新所有与核心数据相关的实例变量。"""
        self.linesnum = len(time_data)
        self.time_con = np.array(time_data)
        self.data_con = np.array(value_data)
        self.vaue = [[self.time_con[i], self.data_con[i]] for i in range(self.linesnum)]
        self.datafit = [''] * self.linesnum

    def _downsample(self, x, y, target_points):
        """使用min-max策略对数据进行降采样以提高绘图性能。"""
        if len(x) <= target_points:
            return x, y
        ratio = int(len(x) / target_points)
        if ratio < 1: ratio = 1
        x_resampled, y_resampled = [], []
        x_resampled.append(x[0])
        y_resampled.append(y[0])
        for i in range(0, len(x), ratio):
            chunk_x = x[i:i+ratio]
            chunk_y = y[i:i+ratio]
            if len(chunk_y) == 0: continue
            min_idx, max_idx = np.argmin(chunk_y), np.argmax(chunk_y)
            if min_idx < max_idx:
                x_resampled.extend([chunk_x[min_idx], chunk_x[max_idx]])
                y_resampled.extend([chunk_y[min_idx], chunk_y[max_idx]])
            elif max_idx < min_idx:
                x_resampled.extend([chunk_x[max_idx], chunk_x[min_idx]])
                y_resampled.extend([chunk_y[max_idx], chunk_y[min_idx]])
            else:
                x_resampled.append(chunk_x[min_idx])
                y_resampled.append(chunk_y[min_idx])
        x_resampled.append(x[-1])
        y_resampled.append(y[-1])
        return np.array(x_resampled), np.array(y_resampled)
  
    # ----------------------------------------------------------------------------------
    # --- 鼠标与UI事件处理 (Mouse and UI Event Handlers) ---
    # ----------------------------------------------------------------------------------
  
    def on_press(self, event):
        """(未使用) 获取鼠标点击在图表上的坐标。"""
        print("my position:", event.button, event.xdata, event.ydata)
        return event.xdata, event.ydata
  
    def oncmask(self, event):
        """用于“区间选择”的鼠标事件处理器。"""
        global stat, leftind, rightind
        ind = np.searchsorted(self.time_con, event.xdata)
        if event.button == 3 and stat == 1: # 状态1: 等待选择左边界
            leftind = ind
            self.a.plot([self.time_con[ind]], [self.data_con[ind]], ".", color="red")
            stat = 2
        elif event.button == 3 and stat == 2: # 状态2: 等待选择右边界
            rightind = ind
            self.a.plot([self.time_con[ind - 1]], [self.data_con[ind - 1]], ".", color="red")
            if leftind > rightind:
                leftind, rightind = rightind - 1, leftind + 1
            self.a.plot(self.time_con[leftind:rightind], self.data_con[leftind:rightind], color="red")
            stat = 3
        elif event.button == 2 and stat == 3: # 状态3: 确认选择
            self.a.plot(self.time_con[leftind:rightind], self.data_con[leftind:rightind], color="blue")
            self.a.plot([self.time_con[leftind]], [self.data_con[leftind]], ".", color="green")
            self.a.plot([self.time_con[rightind]], [self.data_con[rightind]], ".", color="green")
            stat = 4
        self.canvas.draw()

    def on_scroll(self, event):
        """处理鼠标滚轮事件以缩放图表。"""
        if self.toolbar.mode != '': return # 避免与Matplotlib工具栏冲突
        global stat
        if stat in [1, 2, 3]: return # 区间选择模式下禁用
        if event.inaxes != self.a: return
        scale_factor = 1.1
        cur_xlim, cur_ylim = self.a.get_xlim(), self.a.get_ylim()
        xdata, ydata = event.xdata, event.ydata
        if xdata is None or ydata is None: return
        if event.button == 'up': # 放大
            new_width = (cur_xlim[1] - cur_xlim[0]) / scale_factor
            new_height = (cur_ylim[1] - cur_ylim[0]) / scale_factor
        elif event.button == 'down': # 缩小
            new_width = (cur_xlim[1] - cur_xlim[0]) * scale_factor
            new_height = (cur_ylim[1] - cur_ylim[0]) * scale_factor
        else:
            return
        rel_x = (cur_xlim[1] - xdata) / (cur_xlim[1] - cur_xlim[0])
        self.a.set_xlim([xdata - new_width * (1 - rel_x), xdata + new_width * rel_x])

        rel_y = (cur_ylim[1] - ydata) / (cur_ylim[1] - cur_ylim[0])
        self.a.set_ylim([ydata - new_height * (1 - rel_y), ydata + new_height * rel_y])
        self.canvas.draw_idle()

    def on_pan_press(self, event):
        """处理鼠标左键按下事件以开始平移。"""
        if self.toolbar.mode != '': return
        global stat
        if stat in [1, 2, 3] or event.button != 1: return
        if event.inaxes != self.a: return
        self.pan_start_x, self.pan_start_y = event.xdata, event.ydata
        self.pan_pressed = True

    def on_pan_motion(self, event):
        """处理鼠标拖动事件以平移图表。"""
        if self.toolbar.mode != '': return
        if not getattr(self, 'pan_pressed', False): return
        if event.inaxes != self.a or event.xdata is None: return
        dx = event.xdata - self.pan_start_x
        dy = event.ydata - self.pan_start_y
        cur_xlim, cur_ylim = self.a.get_xlim(), self.a.get_ylim()
        self.a.set_xlim([cur_xlim[0] - dx, cur_xlim[1] - dx])
        self.a.set_ylim([cur_ylim[0] - dy, cur_ylim[1] - dy])
        self.canvas.draw_idle()

    def on_pan_release(self, event):
        """处理鼠标释放事件以停止平移。"""
        self.pan_pressed = False
  
    def regionchoose(self):
        """启动“区间选择”模式。"""
        global stat
        stat = 1
        self.canvas.mpl_connect('button_press_event', self.oncmask)
  
    # ----------------------------------------------------------------------------------
    # --- 绘图与数据分析 (Plotting and Data Analysis) ---
    # ----------------------------------------------------------------------------------

    def draw(self):
        """（重新）绘制主数据曲线，大数据量时会自动降采样。"""
        global stat
        stat = 0
        self.a.clear()
        self.a.grid(True)
        self.a.set_xlabel("Time")
        self.a.set_ylabel("Voltage")
        self.a.set_title("ADC-Data")
        threshold = 4000  # 降采样阈值
        if hasattr(self, 'time_con') and self.linesnum > threshold:
            plot_x, plot_y = self._downsample(self.time_con, self.data_con, threshold)
            self.a.plot(plot_x, plot_y, linestyle='-')
        elif hasattr(self, 'time_con'):
            self.a.plot(self.time_con, self.data_con, markersize=3, marker=".")
        self.canvas.draw_idle()
  
    def fitresult(self):
        """计算并绘制线性拟合结果。"""
        fit_x, fit_y = [], []
        if stat == 0: # 对全部数据进行拟合
            aa, bb, self.rr = self.linefit(self.time_con, self.data_con)
            fit_y = aa * self.time_con + bb
            fit_x = self.time_con
        elif stat == 4: # 对选定区间的数据进行拟合
            aa, bb, self.rr = self.linefit(self.time_con[leftind:rightind], self.data_con[leftind:rightind])
            fit_x = self.time_con[leftind:rightind]
            fit_y = aa * fit_x + bb
        elif stat == 5: # 对滤波后的数据进行拟合
            aa, bb, self.rr = self.linefit(filterdata_time, filterdata_data)
            fit_x = np.array(filterdata_time)
            fit_y = aa * fit_x + bb
        else:
            return # 其他状态下不执行

        function_fit = f"y={aa:.3f}*X+{bb:.6f}"
        self.function_fit_export = function_fit
        self.fit_Text.delete(1.0, tk.END)
        self.fit_Text.insert(1.0, function_fit)
        self.fitR_Text.delete(1.0, tk.END)
        self.fitR_Text.insert(1.0, f"{self.rr:.6f}")
        self.a.plot(fit_x, fit_y, linestyle="--")
        self.canvas.draw_idle()
  
    def RMSCAL(self):
        """计算并显示RMS噪声值。"""
        ori_y, fit_y = [], []
        if stat == 0:
            aa, bb, _ = self.linefit(self.time_con, self.data_con)
            ori_y = self.data_con
            fit_y = aa * self.time_con + bb
        elif stat == 4:
            aa, bb, _ = self.linefit(self.time_con[leftind:rightind], self.data_con[leftind:rightind])
            ori_y = self.data_con[leftind:rightind]
            fit_y = aa * self.time_con[leftind:rightind] + bb
        elif stat == 5:
            aa, bb, _ = self.linefit(filterdata_time, filterdata_data)
            ori_y = np.array(filterdata_data)
            fit_y = aa * np.array(filterdata_time) + bb
        else:
            return

        noise = ori_y - fit_y
        vpp_noise = np.max(noise) - np.min(noise)
        sigma_noise = np.std(noise)
    
        self.Noise_Vpp_ex = str(vpp_noise)
        self.Noise_3σ_ex = str(3 * sigma_noise)
        self.Noise_6σ_ex = str(6 * sigma_noise)

        self.Max_Text.delete(1.0, tk.END)
        self.Max_Text.insert(1.0, f"{vpp_noise:.5f}")
        self.Ave_Text.delete(1.0, tk.END)
        self.Ave_Text.insert(1.0, f"{3 * sigma_noise:.5f}")
  
    def linefit(self, x, y):
        """最小二乘法线性拟合。返回斜率a, 截距b, 相关系数r。"""
        if len(x) < 2: return 0, 0, 0
        N = len(x)
        sx, sy = np.sum(x), np.sum(y)
        sxx, syy, sxy = np.sum(x*x), np.sum(y*y), np.sum(x*y)
        a = (sy * sx / N - sxy) / (sx * sx / N - sxx) if (sx * sx / N - sxx) != 0 else 0
        b = (sy - a * sx) / N
        denominator = np.sqrt((sxx - sx*sx/N) * (syy - sy*sy/N))
        r = abs(sy*sx/N - sxy) / denominator if denominator != 0 else 0
        return a, b, r
  
    def setup_config(self):
        """打开ADC参数设置窗口，并将数据转换为真实电压值。"""
        res = self.ask_userinfo()
        if res:
            self.data_con = self.data_con / (2**res[0]) * res[1]
            self.update_data_variables(self.time_con, self.data_con) # 重新同步所有数据变量
            self.tree.delete(*self.tree.get_children())
            for data_tree in self.vaue:
                self.tree.insert('', tk.END, values=data_tree)
            self.draw()
  
    def datafilter_config(self):
        """打开数字滤波设置窗口并应用滤波器。"""
        global stat, filterdata_time, filterdata_data
        stat = 5
        self.draw() # 先重绘原始数据
        filter_window = self.ask_filterinfor()
        if filter_window:
            fwdowidth = int(filter_window[0])
            if fwdowidth < 2 or fwdowidth > (self.linesnum / 10):
                tkinter.messagebox.showwarning(title='Filter Window', message='请输入正确的滤波宽度！')
            else:
                # 使用卷积实现滑动平均滤波，更高效
                filterdata_data = np.convolve(self.data_con, np.ones(fwdowidth)/fwdowidth, mode='valid')
                # 调整时间轴以匹配滤波后数据的长度
                filterdata_time = self.time_con[fwdowidth-1:]
                self.a.plot(filterdata_time, filterdata_data, color="orange")
                self.canvas.draw()
  
    def ask_filterinfor(self):
        inputfilter = digital_filter()
        self.wait_window(inputfilter)
        return inputfilter.filterinfo
  
    def ask_userinfo(self):
        inputDialog = Adc_transfer()
        self.wait_window(inputDialog)
        return inputDialog.userinfo
  
    def open_fft_settings(self):
        """打开FFT设置对话框。"""
        if self.linesnum < 2:
            tkinter.messagebox.showwarning(title='数据不足', message='请先导入或生成有效数据！')
            return
        default_fs = 1 / (self.time_con[1] - self.time_con[0]) if self.linesnum > 1 else 1000
        dialog = FFTSettingDialog(self, default_fs=default_fs)
        self.wait_window(dialog)
        if dialog.settings:
            self.FFT_data(dialog.settings)

    def FFT_data(self, settings):
        """执行FFT分析并绘制频谱图。"""
        global stat
        stat = 0
        Fs = (1 / (self.time_con[1] - self.time_con[0])) if settings["fs_mode"] == "Auto" else settings["fs_value"]
        fre, amp = self.corrected_FFT(Fs, self.data_con, settings["window"])
        y_label = "Amplitude"
        if settings["fft_mode"] == "dB":
            amp = np.maximum(amp, 1e-12) # 避免log(0)
            y_plot = 20 * np.log10(amp)
            y_label = "Amplitude (dB)"
        else:
            y_plot = amp
        self.a.clear()
        self.a.grid(True)
        self.a.set_xlabel("Freq (Hz)")
        self.a.set_ylabel(y_label)
        self.a.set_title(f"FFT Analysis (Window: {settings['window']})")
        self.a.plot(fre, y_plot, markersize=3, marker=".")
        self.canvas.draw()
  
    def corrected_FFT(self, Fs, data, window_name="None (Rectangular)"):
        """改进的FFT计算方法，应用窗函数并进行幅度修正。"""
        L = len(data)
        if L == 0: return [], []
        # 应用窗函数
        window_map = {"Hanning": np.hanning, "Hamming": np.hamming, "Blackman": np.blackman}
        window = window_map.get(window_name, np.ones)(L)
        windowed_data = data * window
        N = int(2 ** np.ceil(np.log2(L))) # 补零到2的幂以提高效率
        fft_y = fft(windowed_data, N)
        correction_factor = np.sum(window)
        if correction_factor == 0: return [], []
        # 计算单边谱
        abs_y = np.abs(fft_y)
        single_side_amp = (abs_y[:N//2 + 1] / correction_factor) * 2
        single_side_amp[0] /= 2 # 直流分量无需乘以2
        fre = np.arange(N//2 + 1) * Fs / N
        return fre, single_side_amp
  
    def export_data(self):
        """将数据和分析结果导出到Excel文件。"""
        if self.linesnum == 0:
            tkinter.messagebox.showwarning(title='无数据', message='没有可导出的数据')
            return
        file_path = filedialog.asksaveasfilename(
                    title='保存文件',
                    filetypes=[("Excel files", "*.xlsx")],
                    defaultextension='.xlsx')
        if not file_path: return
        wb = xlsxwriter.Workbook(file_path)
        style_title = wb.add_format({'bold': True, 'font_name': 'Times New Roman', 'font_size': 12})
        style_data = wb.add_format({'font_name': 'Times New Roman', 'font_size': 12, 'num_format': '0.0000000'})
        ws1 = wb.add_worksheet("All data")
        ws1.write(0, 0, "time", style_title)
        ws1.write(0, 1, "data", style_title)
        for i in range(self.linesnum):
            ws1.write(i + 1, 0, self.time_con[i], style_data)
            ws1.write(i + 1, 1, self.data_con[i], style_data)
        if stat == 4: # 如果有选区，额外保存选区数据和分析结果
            ws2 = wb.add_worksheet("Selected Data")
            ws2.write(0, 0, "time", style_title)
            ws2.write(0, 1, "data", style_title)
            for i in range(leftind, rightind):
                ws2.write(i - leftind + 1, 0, self.time_con[i], style_data)
                ws2.write(i - leftind + 1, 1, self.data_con[i], style_data)
            ws3 = wb.add_worksheet("Calculating Results")
            ws3.write(0, 0, "Noise_Vpp:", style_title)
            ws3.write(0, 1, self.Noise_Vpp_ex, style_data)
            ws3.write(1, 0, "Noise_3σ:", style_title)
            ws3.write(1, 1, self.Noise_3σ_ex, style_data)
            ws3.write(2, 0, "Noise_6σ:", style_title)
            ws3.write(2, 1, self.Noise_6σ_ex, style_data)
            ws3.write(3, 0, "Line Fitting:", style_title)
            ws3.write(3, 1, self.function_fit_export, style_data)
        wb.close()
        tkinter.messagebox.showinfo(title='提示', message='数据导出成功')

    # ----------------------------------------------------------------------------------
    # --- 表格编辑功能 (Table Editing Features) ---
    # ----------------------------------------------------------------------------------

    def add_table_row(self):
        """在数据表末尾添加一个新行。"""
        new_time = 0.0
        if self.linesnum > 1:
            avg_interval = (self.time_con[-1] - self.time_con[0]) / (self.linesnum - 1)
            new_time = self.time_con[-1] + avg_interval
        elif self.linesnum == 1:
            new_time = self.time_con[-1] + 1.0
        new_time = round(new_time, 6)
        self.time_con = np.append(self.time_con, new_time)
        self.data_con = np.append(self.data_con, 0.0)
        self.vaue.append([new_time, 0.0])
        self.linesnum += 1
        new_item = self.tree.insert('', tk.END, values=(new_time, 0.0))
        self.tree.yview_moveto(1)
        self.tree.selection_set(new_item)
        self.draw()

    def delete_table_row(self):
        """删除在表格中选中的行。"""
        selected_items = self.tree.selection()
        if not selected_items:
            tkinter.messagebox.showinfo("提示", "请先在表格中选择要删除的行")
            return
        indices_to_remove = [self.tree.index(item) for item in selected_items]
        for item in selected_items:
            self.tree.delete(item)
        self.time_con = np.delete(self.time_con, indices_to_remove)
        self.data_con = np.delete(self.data_con, indices_to_remove)
        self.linesnum = len(self.time_con)
        self.vaue = [[self.time_con[i], self.data_con[i]] for i in range(self.linesnum)]
        self.draw()

    def paste_from_clipboard(self):
        """从剪贴板粘贴数据并覆盖现有数据。"""
        try:
            clipboard_data = self.clipboard_get()
        except tk.TclError:
            tkinter.messagebox.showwarning("粘贴错误", "剪贴板为空或不包含文本数据。")
            return
        if self.linesnum > 0 and not tkinter.messagebox.askyesno("确认操作", "粘贴数据将覆盖当前所有数据，是否继续？"):
            return
        self.clear_all_data()
        new_time, new_data = [], []
        lines = clipboard_data.strip().split('\n')
        for i, line in enumerate(lines):
            try:
                parts = line.replace('\t', ' ').split()
                if len(parts) >= 2:
                    new_time.append(float(parts[0]))
                    new_data.append(float(parts[1]))
            except (ValueError, IndexError):
                tkinter.messagebox.showwarning("解析警告", f"第 {i+1} 行数据格式不正确，已跳过。")
                continue
        if not new_time:
            tkinter.messagebox.showerror("粘贴失败", "未能从剪贴板中解析出任何有效数据。")
            return
        self.update_data_variables(new_time, new_data)
        for item in self.vaue:
            self.tree.insert('', tk.END, values=item)
        self.draw()
        tkinter.messagebox.showinfo("成功", f"成功从剪贴板粘贴 {self.linesnum} 条数据。")

    def on_tree_double_click(self, event):
        """处理双击表格单元格事件，创建临时编辑框。"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell": return
        column_id = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        if not item_id: return
        x, y, width, height = self.tree.bbox(item_id, column_id)
        value = self.tree.set(item_id, column_id)
        entry = ttk.Entry(self.tree)
        entry.place(x=x, y=y, width=width, height=height)
        entry.insert(0, value)
        entry.focus_force()
        entry.bind("<Return>", lambda e: self._save_tree_edit(entry, item_id, column_id))
        entry.bind("<FocusOut>", lambda e: self._save_tree_edit(entry, item_id, column_id))

    def _save_tree_edit(self, entry, item_id, column_id):
        """保存编辑框中的值，更新数据和UI。"""
        if not entry.winfo_exists(): return
        try:
            new_value = float(entry.get())
        except ValueError:
            entry.destroy()
            return
        self.tree.set(item_id, column=column_id, value=new_value)
        item_index = self.tree.index(item_id)
        column_index = int(column_id.replace('#', '')) - 1
        self.vaue[item_index][column_index] = new_value
        if column_index == 0:
            self.time_con[item_index] = new_value
        else:
            self.data_con[item_index] = new_value
        entry.destroy()
        self.draw()

# =====================================================================================
# --- 程序主入口 (Main Execution Block) ---
# =====================================================================================

if __name__ == '__main__':
    # 全局变量定义
    # stat: 程序状态标志，用于控制不同的鼠标交互模式
    # 0: 默认状态 (绘图/平移/缩放)
    # 1: 区间选择模式 - 等待选择左边界
    # 2: 区间选择模式 - 等待选择右边界
    # 3: 区间选择模式 - 确认选择
    # 4: 区间选择完成
    # 5: 数字滤波模式
    stat = 0
    leftind, rightind = 0, 0  # 用于存储区间选择的左右索引
    filterdata_time, filterdata_data = [], [] # 用于存储滤波后的数据
  
    # 实例化并运行主应用
    ZMJ_PORTAL = MY_GUI()
    ZMJ_PORTAL.mainloop()